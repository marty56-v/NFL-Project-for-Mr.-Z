import requests
from bs4 import BeautifulSoup
import numpy as np
from openpyxl import *
from openpyxl.styles import Font
import getpass



#Creates an empty array
qbStats = np.empty(shape=(6,40),dtype='object') 
#stuff
zLoopCount = 0
fantPts = 0
week = input("Stats from week:\n")
wb = Workbook()
ws = wb.active

# https://www.nfl.com/stats/player-stats/category/passing/2020/REG/ALL/passingyards/desc
#testing url https://codeprojects.org/T009Fzzrr8XXzzqn-K9ig1KIrT8LQORUNJs9FeEBWhU/
#URL = "https://www.nfl.com/stats/player-stats/category/rushing/2020/REG/all/rushingyards/desc"


#gets all the stats for the qb plus the names
def qbStatsGet (weekNum = 1):
  counter = 0
  counter2 = 0
  counter3 = 0
  counter4 = 0
  counter5 = 0
  results = 0 
  results2 = None
  results3 =  0
  results4 =  0
  results5 =  0
  
  #set to week 2 right now
  qbURL = "https://www.pro-football-reference.com/play-index/pgl_finder.cgi?request=1&match=game&year_min=2020&year_max=2020&season_start=1&season_end=-1&pos%5B%5D=QB&is_starter=E&game_num_min=0&game_num_max=99&week_num_min=" + str(week) + "&week_num_max=" + str(week) + "&c1stat=pass_att&c1comp=gt&c1val=1&c2stat=rush_yds&c2comp=gt&c5val=1.0&order_by=pass_rating"
  page = requests.get(qbURL)
  soup = BeautifulSoup(page.content, 'html.parser')
  
  
  #Reminder for me to not use classes if they are not needed
  #gets player names  
  for i in soup.find_all("td", class_ = "left", attrs ={"data-stat": "player"}):
    results2 = i.getText()
    qbStats[0,counter2] = results2
    counter2 += 1


    pass
  #gets passing yards 
  for i in soup.find_all("td", attrs={"data-stat": "pass_yds"}):
    results = i.getText()
    qbStats[1,counter] = int(results)
    counter += 1
    pass
  #gets passing touchdowns
  for i in soup.find_all("td", attrs={"data-stat": "pass_td"}):
    results3 = i.getText()
    qbStats[2,counter3] = int(results3)
    counter3 += 1
    pass
  #gets rushing yards
  for i in soup.find_all("td", attrs={"data-stat": "rush_yds"}):
    results4 = i.getText()
    qbStats[3,counter4] = int(results4)
    counter4 += 1
    pass
  #gets rushing touchdowns
  for i in soup.find_all("td",attrs = {"data-stat" : "rush_td"}):
    results5 =  i.getText()
    qbStats[4,counter5] = int(results5)
    counter5 += 1
    pass


qbStatsGet(week)

#prints some stuff
while zLoopCount < 40:
  if (qbStats[0,zLoopCount] == None):
    break
  else:
    colNumName = 1 #identifies column for Qb name
    colNumPassYds = 2 #identifies column for pass yds
    colNumPassTDs = 3 #identifies column for pass tds
    colNumRushYds = 4 #identifies column for rush yds
    colNumRushTDs = 5 #identifies column for rush tds
    colNumFantPts = 6 #identifies column for fant pts
    TGREEN =  '\033[92m' # Green-ish Text
    TWHITE = '\033[37m' # White-ish Text
    TBLUE = '\033[94m' #blue text
    print(TGREEN + qbStats[0,zLoopCount], TWHITE)

    #fantasy pts calc
    fantPts = qbStats[1,zLoopCount]/25 + qbStats[2,zLoopCount] * 6 + qbStats[3,zLoopCount]/10 + qbStats[4,zLoopCount] * 6
    #round number to 1/100th
    fantPts = round(fantPts, 2)
    #converted qbStats to a string in order to concatenate
    #print stats
    print(str(qbStats[1,zLoopCount]) + " pass yds")

    print(str(qbStats[2,zLoopCount])+" pass TDs")
    print(str(qbStats[3,zLoopCount])+" rush yds")
    print(str(qbStats[4,zLoopCount])+ " rush TDs")
    print(TBLUE +str(fantPts)+ " Total Pts", TGREEN)
    qbStats[5,zLoopCount] = fantPts


    #puts the stats into the spreadsheet
    ws.cell(row=zLoopCount+2, column=colNumName).value=qbStats[0,zLoopCount]
    ws.cell(row=zLoopCount+2, column=colNumPassYds).value=qbStats[1,zLoopCount]
    ws.cell(row = zLoopCount + 2, column = colNumPassTDs).value = qbStats[2,zLoopCount]   
    ws.cell(row = zLoopCount + 2, column = colNumRushYds).value = qbStats[3,zLoopCount]
    ws.cell(row = zLoopCount + 2, column = colNumRushTDs).value = qbStats[4,zLoopCount]    
    ws.cell(row = zLoopCount+ 2, column = colNumFantPts).value = qbStats[5,zLoopCount]

    print("")

  zLoopCount += 1

  pass

#Headings for the columns
fontObj = Font(name='Arial', size=12, bold=True)
ws['A1'].font = fontObj
ws["A1"] = "Name:"  
ws['B1'].font = fontObj
ws["B1"] = "Pass Yds:"  
ws['C1'].font = fontObj
ws['C1'] = "Pass TDs:"
ws['D1'].font = fontObj
ws['D1'] = "Rush Yds:"
ws['E1'].font = fontObj
ws['E1'] = "Rush TDs:"
ws['F1'].font = fontObj
ws['F1'] = "Fant Pts:"

#Column width sizes should all be 25 and if you use any extra columns make them 25 (if you REALLY need to increase the width then change them all to be the same)
ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 25
ws.column_dimensions["C"].width = 25
ws.column_dimensions["D"].width = 25
ws.column_dimensions["E"].width = 25
ws.column_dimensions["F"].width = 25

##### SOMETHING REALLY IMPORTANT TO KNOW IS THAT REPL WILL NOT UPDATE THE SPREADSHEET UNLESS YOU RELOAD AND RUN THE PROGRAM OTHERWISE DOWNLOADING THE SPREADSHEET WILL NOT WORK!
#wb.save("nflFantasy.xlsx")
wb.save('/home/'+getpass.getuser()+'/Desktop/nflFantasy.xlsx')

